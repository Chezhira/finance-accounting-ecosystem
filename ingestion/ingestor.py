"""
ingestion/ingestor.py — FinOps Ecosystem Data Ingestor
Phase 1: Email, file upload, webhook, raw text
Phase 5: OCR fallback for image-rendered PDFs (Stripe, scanned docs, etc.)

OCR Strategy:
  1. Try pdfplumber first (fast, text-layer PDFs)
  2. If extracted text is below MIN_TEXT_THRESHOLD characters → classify as image-based PDF
  3. Rasterize pages via pdf2image (poppler backend)
  4. Run tesseract OCR on each page image
  5. Concatenate page text with page markers
  6. Return OCR result as if it were normal file text

Requirements (add to requirements.txt):
  pdfplumber>=0.10.0
  pytesseract>=0.3.10
  pdf2image>=1.17.0
  Pillow>=10.0.0
  (system: tesseract-ocr, poppler-utils)
"""

import io
import os
import json
import logging
import email as email_lib
from email import policy as email_policy
from typing import Optional
from pathlib import Path

import pdfplumber

# ── OCR dependencies (graceful fallback if not installed) ──────────────────────
try:
    import pytesseract
    from pdf2image import convert_from_bytes
    from PIL import Image
    OCR_AVAILABLE = True
except ImportError:
    OCR_AVAILABLE = False

# ── Optional spreadsheet parsing ───────────────────────────────────────────────
try:
    import csv
    CSV_AVAILABLE = True
except ImportError:
    CSV_AVAILABLE = False

try:
    import openpyxl
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False

logger = logging.getLogger(__name__)

# ──────────────────────────────────────────────────────────────────────────────
# Constants
# ──────────────────────────────────────────────────────────────────────────────

# If pdfplumber extracts fewer characters than this, treat PDF as image-based
MIN_TEXT_THRESHOLD = 80

# OCR config — single-column docs; optimise for financial documents
TESSERACT_CONFIG = r"--oem 3 --psm 6"

# DPI for PDF rasterisation — 300 gives good OCR accuracy; 200 is faster
OCR_DPI = 300


# ──────────────────────────────────────────────────────────────────────────────
# Ingestion result dataclass (plain dict for simplicity)
# ──────────────────────────────────────────────────────────────────────────────

def _make_result(
    text: str,
    source_type: str,
    filename: Optional[str] = None,
    metadata: Optional[dict] = None,
    ocr_used: bool = False,
    ocr_page_count: int = 0,
    ocr_confidence: Optional[float] = None,
    warnings: Optional[list] = None,
) -> dict:
    return {
        "text": text.strip(),
        "source_type": source_type,         # "file_pdf" | "file_csv" | "file_xlsx" | "email" | "webhook" | "raw_text"
        "filename": filename,
        "metadata": metadata or {},
        "ocr_used": ocr_used,
        "ocr_page_count": ocr_page_count,
        "ocr_confidence": ocr_confidence,
        "char_count": len(text.strip()),
        "warnings": warnings or [],
    }


# ──────────────────────────────────────────────────────────────────────────────
# Core PDF Extraction
# ──────────────────────────────────────────────────────────────────────────────

def _extract_pdf_text(content: bytes) -> str:
    """
    Extract text from PDF using pdfplumber.
    Returns concatenated text from all pages.
    Returns empty string if no text layer found.
    """
    text_parts = []
    try:
        with pdfplumber.open(io.BytesIO(content)) as pdf:
            for i, page in enumerate(pdf.pages, start=1):
                page_text = page.extract_text() or ""
                if page_text.strip():
                    text_parts.append(f"[Page {i}]\n{page_text}")
    except Exception as e:
        logger.warning(f"pdfplumber extraction failed: {e}")
    return "\n\n".join(text_parts)


def _is_image_based_pdf(extracted_text: str) -> bool:
    """
    Heuristic: if pdfplumber returned very little text, the PDF is image-based.
    """
    return len(extracted_text.strip()) < MIN_TEXT_THRESHOLD


def _ocr_pdf(content: bytes) -> tuple[str, int, Optional[float]]:
    """
    Rasterise PDF pages and run Tesseract OCR on each.
    Returns (combined_text, page_count, avg_confidence).
    Raises RuntimeError if OCR is not available.
    """
    if not OCR_AVAILABLE:
        raise RuntimeError(
            "OCR libraries not installed. Run: "
            "pip install pytesseract pdf2image Pillow "
            "and install system packages: tesseract-ocr poppler-utils"
        )

    logger.info("Starting OCR processing — rasterising PDF pages at %d DPI", OCR_DPI)

    try:
        images = convert_from_bytes(content, dpi=OCR_DPI)
    except Exception as e:
        raise RuntimeError(f"pdf2image conversion failed: {e}") from e

    page_texts = []
    confidences = []

    for i, img in enumerate(images, start=1):
        try:
            # Get both text and confidence data
            ocr_data = pytesseract.image_to_data(
                img,
                config=TESSERACT_CONFIG,
                output_type=pytesseract.Output.DICT,
            )

            # Filter out low-confidence or empty words
            words = []
            for j, word in enumerate(ocr_data["text"]):
                conf = int(ocr_data["conf"][j])
                if conf > 0 and word.strip():
                    words.append(word)
                    confidences.append(conf)

            page_text = " ".join(words)
            if page_text.strip():
                page_texts.append(f"[Page {i} — OCR]\n{page_text}")

            logger.debug(
                "Page %d OCR: %d words extracted",
                i, len(words)
            )

        except Exception as e:
            logger.warning("OCR failed on page %d: %s", i, e)
            page_texts.append(f"[Page {i} — OCR FAILED: {e}]")

    combined = "\n\n".join(page_texts)
    page_count = len(images)
    avg_confidence = (
        round(sum(confidences) / len(confidences), 1)
        if confidences else None
    )

    logger.info(
        "OCR complete: %d pages, avg confidence %.1f%%",
        page_count,
        avg_confidence or 0,
    )

    return combined, page_count, avg_confidence


# ──────────────────────────────────────────────────────────────────────────────
# CSV / Excel helpers
# ──────────────────────────────────────────────────────────────────────────────

def _parse_csv(content: bytes) -> str:
    """Parse CSV bytes and return a plain-text table representation."""
    try:
        text = content.decode("utf-8", errors="replace")
        reader = csv.reader(io.StringIO(text))
        rows = list(reader)
        if not rows:
            return ""
        # Format as aligned pipe-delimited text for the agent
        col_widths = [
            max(len(str(row[i])) for row in rows if i < len(row))
            for i in range(len(rows[0]))
        ]
        lines = []
        for row in rows:
            padded = [
                str(row[i]).ljust(col_widths[i]) if i < len(row) else " " * col_widths[i]
                for i in range(len(col_widths))
            ]
            lines.append("| " + " | ".join(padded) + " |")
        return "\n".join(lines)
    except Exception as e:
        logger.warning("CSV parsing failed: %s", e)
        return content.decode("utf-8", errors="replace")


def _parse_xlsx(content: bytes) -> str:
    """Parse Excel bytes and return a plain-text multi-sheet representation."""
    if not XLSX_AVAILABLE:
        return "[XLSX parsing unavailable — install openpyxl]"
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        sections = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                # Skip entirely empty rows
                if any(cell is not None for cell in row):
                    rows.append("\t".join(str(c) if c is not None else "" for c in row))
            if rows:
                sections.append(f"[Sheet: {sheet_name}]\n" + "\n".join(rows))
        return "\n\n".join(sections)
    except Exception as e:
        logger.warning("XLSX parsing failed: %s", e)
        return f"[XLSX parse error: {e}]"


# ──────────────────────────────────────────────────────────────────────────────
# DataIngestor — main public class
# ──────────────────────────────────────────────────────────────────────────────

class DataIngestor:
    """
    Single entry point for all data ingestion in the FinOps ecosystem.

    Usage:
        result = DataIngestor.from_file(content_bytes, "invoice.pdf")
        result = DataIngestor.from_raw_text("Revenue for Q1 was TZS 5,000,000")
        result = DataIngestor.from_email(raw_email_string)
        result = DataIngestor.from_webhook(payload_dict, "quickbooks")

    All methods return a dict with keys:
        text          — extracted/combined text ready for agent
        source_type   — origin of data
        filename      — original filename (if applicable)
        metadata      — extra parsed info (headers, sender, system etc.)
        ocr_used      — bool — was OCR applied?
        ocr_page_count— int  — number of pages OCR'd (0 if not used)
        ocr_confidence— float|None — average Tesseract confidence %
        char_count    — int — length of extracted text
        warnings      — list of non-fatal issues
    """

    # ── File ingestion ──────────────────────────────────────────────────────

    @staticmethod
    def from_file(content: bytes, filename: str) -> dict:
        """
        Ingest a file upload.

        Supports: PDF, CSV, XLSX/XLS, JSON, TXT and plain text fallback.

        CRITICAL: Pass raw bytes — never decode to string first.
                  PDF binary is destroyed by decode(). pdfplumber handles internally.
        """
        ext = Path(filename).suffix.lower()
        warnings = []

        # ── PDF ─────────────────────────────────────────────────────────────
        if ext == ".pdf":
            text = _extract_pdf_text(content)

            if _is_image_based_pdf(text):
                logger.info(
                    "PDF '%s' has no text layer (%d chars) — attempting OCR",
                    filename, len(text.strip())
                )
                warnings.append(
                    "PDF appears to be image-based (e.g. scanned or Stripe-generated). "
                    "OCR was applied — verify extracted text for accuracy."
                )
                try:
                    ocr_text, page_count, confidence = _ocr_pdf(content)
                    if ocr_text.strip():
                        return _make_result(
                            text=ocr_text,
                            source_type="file_pdf_ocr",
                            filename=filename,
                            metadata={"pages_ocr'd": page_count},
                            ocr_used=True,
                            ocr_page_count=page_count,
                            ocr_confidence=confidence,
                            warnings=warnings,
                        )
                    else:
                        warnings.append(
                            "OCR returned no text. PDF may be blank, heavily degraded, "
                            "or in an unsupported format. Use Raw Text tab to paste content manually."
                        )
                        return _make_result(
                            text="[OCR returned no text — see warnings]",
                            source_type="file_pdf_ocr",
                            filename=filename,
                            ocr_used=True,
                            ocr_page_count=page_count,
                            ocr_confidence=confidence,
                            warnings=warnings,
                        )
                except RuntimeError as e:
                    warnings.append(f"OCR unavailable: {e}")
                    logger.error("OCR error for '%s': %s", filename, e)
                    return _make_result(
                        text=(
                            text if text.strip()
                            else "[PDF has no text layer and OCR is not available — install pytesseract + poppler]"
                        ),
                        source_type="file_pdf",
                        filename=filename,
                        warnings=warnings,
                    )

            # Normal text-layer PDF
            return _make_result(
                text=text,
                source_type="file_pdf",
                filename=filename,
            )

        # ── CSV ─────────────────────────────────────────────────────────────
        elif ext == ".csv":
            text = _parse_csv(content)
            return _make_result(
                text=text,
                source_type="file_csv",
                filename=filename,
            )

        # ── Excel ────────────────────────────────────────────────────────────
        elif ext in (".xlsx", ".xls"):
            text = _parse_xlsx(content)
            return _make_result(
                text=text,
                source_type="file_xlsx",
                filename=filename,
            )

        # ── JSON ─────────────────────────────────────────────────────────────
        elif ext == ".json":
            try:
                parsed = json.loads(content.decode("utf-8", errors="replace"))
                text = json.dumps(parsed, indent=2)
            except Exception:
                text = content.decode("utf-8", errors="replace")
            return _make_result(
                text=text,
                source_type="file_json",
                filename=filename,
            )

        # ── Plain text / fallback ─────────────────────────────────────────────
        else:
            text = content.decode("utf-8", errors="replace")
            return _make_result(
                text=text,
                source_type="file_txt",
                filename=filename,
            )

    # ── Raw text ingestion ──────────────────────────────────────────────────

    @staticmethod
    def from_raw_text(text: str) -> dict:
        """
        Ingest plain text pasted directly by the user or operator.
        """
        if not text or not text.strip():
            return _make_result(
                text="[Empty input]",
                source_type="raw_text",
                warnings=["Input was empty or whitespace only"],
            )
        return _make_result(
            text=text,
            source_type="raw_text",
        )

    # ── Email ingestion ─────────────────────────────────────────────────────

    @staticmethod
    def from_email(raw_email) -> dict:
        """
        Ingest an email.

        Accepts:
          - raw email string (RFC 2822 format)
          - dict with keys: from, to, subject, body (and optionally attachments)
        """
        metadata = {}
        text_parts = []
        warnings = []

        if isinstance(raw_email, dict):
            # Pre-parsed dict input
            metadata = {
                "from": raw_email.get("from", ""),
                "to": raw_email.get("to", ""),
                "subject": raw_email.get("subject", ""),
                "date": raw_email.get("date", ""),
            }
            body = raw_email.get("body", "") or raw_email.get("text", "") or raw_email.get("html", "")
            text_parts.append(f"Subject: {metadata['subject']}")
            text_parts.append(f"From: {metadata['from']}")
            text_parts.append(f"Body:\n{body}")

            # Handle inline attachments if provided as list of {"filename":..., "content": bytes}
            attachments = raw_email.get("attachments", [])
            for att in attachments:
                att_filename = att.get("filename", "attachment")
                att_content = att.get("content", b"")
                if isinstance(att_content, bytes) and att_content:
                    att_result = DataIngestor.from_file(att_content, att_filename)
                    text_parts.append(
                        f"\n[Attachment: {att_filename}]\n{att_result['text']}"
                    )
                    if att_result.get("ocr_used"):
                        warnings.append(
                            f"Attachment '{att_filename}' required OCR — "
                            f"confidence: {att_result.get('ocr_confidence')}%"
                        )

        elif isinstance(raw_email, str):
            # Raw RFC 2822 email string
            try:
                msg = email_lib.message_from_string(raw_email, policy=email_policy.default)
                metadata = {
                    "from": str(msg.get("From", "")),
                    "to": str(msg.get("To", "")),
                    "subject": str(msg.get("Subject", "")),
                    "date": str(msg.get("Date", "")),
                }
                text_parts.append(f"Subject: {metadata['subject']}")
                text_parts.append(f"From: {metadata['from']}")

                # Walk MIME parts
                for part in msg.walk():
                    ctype = part.get_content_type()
                    disposition = str(part.get("Content-Disposition", ""))

                    if "attachment" in disposition:
                        att_filename = part.get_filename() or "attachment"
                        att_bytes = part.get_payload(decode=True)
                        if att_bytes:
                            att_result = DataIngestor.from_file(att_bytes, att_filename)
                            text_parts.append(
                                f"\n[Attachment: {att_filename}]\n{att_result['text']}"
                            )
                            if att_result.get("ocr_used"):
                                warnings.append(
                                    f"Attachment '{att_filename}' required OCR — "
                                    f"confidence: {att_result.get('ocr_confidence')}%"
                                )
                    elif ctype == "text/plain":
                        body = part.get_payload(decode=True)
                        if body:
                            text_parts.append(
                                "Body:\n" + body.decode("utf-8", errors="replace")
                            )
                    elif ctype == "text/html" and len(text_parts) < 3:
                        # Only use HTML if no plain text found
                        body = part.get_payload(decode=True)
                        if body:
                            text_parts.append(
                                "Body (HTML):\n" + body.decode("utf-8", errors="replace")
                            )

            except Exception as e:
                logger.warning("Email parse failed: %s", e)
                warnings.append(f"Email parse error: {e}")
                text_parts.append(raw_email)

        else:
            warnings.append(f"Unexpected email input type: {type(raw_email)}")
            text_parts.append(str(raw_email))

        return _make_result(
            text="\n\n".join(text_parts),
            source_type="email",
            metadata=metadata,
            warnings=warnings,
        )

    # ── Webhook ingestion ───────────────────────────────────────────────────

    @staticmethod
    def from_webhook(payload: dict, system: str) -> dict:
        """
        Ingest a webhook payload from an accounting system.

        Supported systems: quickbooks, fishbowl, bill, stripe, xero, odoo, generic
        """
        system_lower = system.lower().replace(".", "").replace("-", "")
        metadata = {"system": system, "event_type": payload.get("eventType", "")}
        text_parts = [f"Webhook from: {system}"]

        if "eventType" in payload:
            text_parts.append(f"Event: {payload['eventType']}")

        # System-specific extraction
        if "quickbooks" in system_lower:
            text_parts.append(_format_qb_payload(payload))

        elif "fishbowl" in system_lower:
            text_parts.append(_format_fishbowl_payload(payload))

        elif "bill" in system_lower:
            text_parts.append(_format_bill_payload(payload))

        elif "stripe" in system_lower:
            text_parts.append(_format_stripe_payload(payload))

        elif "xero" in system_lower:
            text_parts.append(_format_xero_payload(payload))

        else:
            # Generic: pretty-print the whole payload
            text_parts.append(json.dumps(payload, indent=2, default=str))

        return _make_result(
            text="\n\n".join(text_parts),
            source_type="webhook",
            metadata=metadata,
        )


# ──────────────────────────────────────────────────────────────────────────────
# Webhook formatters — extract human-readable summaries per system
# ──────────────────────────────────────────────────────────────────────────────

def _format_qb_payload(payload: dict) -> str:
    lines = []
    for entity in payload.get("entities", []):
        lines.append(
            f"Entity: {entity.get('name')} | ID: {entity.get('id')} | "
            f"Operation: {entity.get('operation')}"
        )
    if payload.get("Bill"):
        b = payload["Bill"]
        lines.append(
            f"Bill #{b.get('DocNumber', 'N/A')}: "
            f"Vendor {b.get('VendorRef', {}).get('name', 'N/A')} | "
            f"Total {b.get('CurrencyRef', {}).get('value', '')} "
            f"{b.get('TotalAmt', 'N/A')}"
        )
    return "\n".join(lines) if lines else json.dumps(payload, indent=2, default=str)


def _format_fishbowl_payload(payload: dict) -> str:
    lines = []
    order_type = payload.get("OrderType", payload.get("type", "Order"))
    lines.append(f"Type: {order_type}")
    for key in ("OrderNum", "PONum", "SONum", "CustomerName", "VendorName",
                "TotalAmount", "Currency", "Status", "DateCreated"):
        if key in payload:
            lines.append(f"{key}: {payload[key]}")
    items = payload.get("Items", payload.get("LineItems", []))
    if items:
        lines.append("Line Items:")
        for item in items:
            lines.append(
                f"  - {item.get('PartNum', 'N/A')} | Qty: {item.get('Qty', 'N/A')} | "
                f"Unit Price: {item.get('UnitCost', item.get('UnitPrice', 'N/A'))}"
            )
    return "\n".join(lines) if lines else json.dumps(payload, indent=2, default=str)


def _format_bill_payload(payload: dict) -> str:
    lines = []
    for key in ("id", "entity", "amount", "currency", "dueDate",
                "invoiceDate", "vendorName", "status", "description"):
        if key in payload:
            lines.append(f"{key}: {payload[key]}")
    line_items = payload.get("billLineItems", [])
    if line_items:
        lines.append("Line Items:")
        for li in line_items:
            lines.append(
                f"  - {li.get('description', 'N/A')}: "
                f"{li.get('amount', 'N/A')} {li.get('chartOfAccountId', '')}"
            )
    return "\n".join(lines) if lines else json.dumps(payload, indent=2, default=str)


def _format_stripe_payload(payload: dict) -> str:
    lines = []
    event_type = payload.get("type", "")
    lines.append(f"Event: {event_type}")
    data_obj = payload.get("data", {}).get("object", {})
    if data_obj:
        for key in ("id", "amount", "currency", "status", "description",
                    "customer", "created", "invoice", "payment_intent"):
            if key in data_obj:
                val = data_obj[key]
                # Convert Stripe cents to currency units
                if key == "amount" and isinstance(val, int):
                    val = f"{val / 100:.2f}"
                lines.append(f"{key}: {val}")
    return "\n".join(lines) if lines else json.dumps(payload, indent=2, default=str)


def _format_xero_payload(payload: dict) -> str:
    lines = []
    for event in payload.get("events", []):
        lines.append(
            f"Event: {event.get('eventType')} | "
            f"Resource: {event.get('resourceType')} | "
            f"ID: {event.get('resourceId')}"
        )
    return "\n".join(lines) if lines else json.dumps(payload, indent=2, default=str)
